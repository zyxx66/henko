import openpyxl

file = 'C:/Users/zyxx/Desktop/sumup_2.xlsx'

wb = openpyxl.load_workbook(file)
ws = wb.active

cell_row = 5
k = 1
while True:
    if ws.cell(1, cell_row).value != None:
        cell_row += 3
        k += 1
    else:
        print('k=%d'%k)
        break

for i in range(24):
    ws.cell(29, 3 * i + 1).value = str(650 + i * 50)
    for k in range(k):
        # ws.cell(30,2).value = ws.cell(3,2).value
        # ws.cell(30,3).value = ws.cell(3,3).value
        #
        # ws.cell(31,2).value = ws.cell(3,5).value
        # ws.cell(31,3).value = ws.cell(3,6).value
        #
        # ws.cell(30,5).value = ws.cell(4,2).value
        # ws.cell(30,6).value = ws.cell(4,3).value

        ws.cell(30 + k, 2 + i * 3).value = ws.cell(3 + i, 2 + k * 3).value
        ws.cell(30 + k, 3 + i * 3).value = ws.cell(3 + i, 3 + 3 * k).value

wb.save(file)
