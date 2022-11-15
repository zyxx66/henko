import openpyxl
import create_graph
file = 'C:/Users/zyxx/Desktop/sumup.xlsx'
# wb = openpyxl.load_workbook(file)
# ws = wb.active

# 3,2 --> 28,2  |   3,3 --> 28,3
# 3,6 --> 29,2  |   3,7 --> 29,3
# 4,2 --> 28,6  |   4,3 --> 28,7
# 4,6 --> 29,6  |   4,7 --> 29,7
for i in range(8):
    if False:
        for k in  range(22):
            ws.cell(28+i,2+4*k).value = ws.cell(3+k,2+i*4).value
            ws.cell(28+i,3+4*k).value = ws.cell(3+k,3+i*4).value
for i in range(22):
    create_graph.create_scatter(file,'Sheet1','雲量(%)','偏光度(%)','','B'+str(i*10+35),[3+4*i,28,3+4*i,35,2+4*i,28,2+4*i,35])

# wb.save(file)