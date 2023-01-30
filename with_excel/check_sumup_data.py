# 毎日のデータをまとめたファイルをグラフ化するプログラムである
# 注意　11月3日前の毎日データファイルは使用不可
import csv
import openpyxl

folder = 'C:/Users/zyxx/Desktop/test_csv/daily_data/result/'
csv_file = 'sumup.csv'
xlsx_file = 'sumup.xlsx'

xlsx_column = {}
# 時間別でデータの列目を記録
# xlsx_column = {'600':[2,3,4],'630':[5,6,7]……}のような辞書を作る
for i in range(26):
    if i % 2 == 0:
        if 600 + i // 2 * 100 < 1000:
            xlsx_column['0' + str(600 + i // 2 * 100)] = [i * 3 + 2, i * 3 + 3, i * 3 + 4]
        else:
            xlsx_column[str(600 + i // 2 * 100)] = [i * 3 + 2, i * 3 + 3, i * 3 + 4]
    else:
        if 630 +i //2*100 <1000:
            xlsx_column['0'+str(630 + i // 2 * 100)] = [i * 3 + 2, i * 3 + 3, i * 3 + 4]
        else:
            xlsx_column[str(630 + i // 2 * 100)] = [i * 3 + 2, i * 3 + 3, i * 3 + 4]


print(xlsx_column)
work_book = openpyxl.Workbook()
work_sheet = work_book.active
sum_up_list = []

# ------------1列目と2列目に書き込むやつ-----------------
align = openpyxl.styles.Alignment(horizontal='center', vertical='center')
for i in range(26):
    if i % 2 == 0:
        work_sheet.cell(1, i * 3 + 2).value = str(i // 2 + 6) + ':00'
    else:
        work_sheet.cell(1, i * 3 + 2).value = str(i // 2 + 6) + ':30'
    work_sheet.cell(2, i * 3 + 2).value = '偏光度(%)'
    work_sheet.cell(2, i * 3 + 3).value = '雲量(%)'
    work_sheet.cell(2, i * 3 + 4).value = 'ch1'
    for j in range(2, 5):
        work_sheet.cell(2, i * 3 + j).alignment = align
    work_sheet.cell(1, i * 3 + 2).alignment = align
    work_sheet.merge_cells(start_row=1, start_column=i * 3 + 2, end_row=1, end_column=i * 3 + 4)

sumup_row = 2

with open(folder + csv_file) as open_csv_file:
    for row in csv.reader(open_csv_file):
        sum_up_list.append(row)
print(sum_up_list)
# xlsxファイルにデータを入力
for data in sum_up_list:
    # もし時間が出現したら
    if '-' in data[0]:
        sumup_row += 1
        work_sheet.cell(sumup_row, 1).value = data[0]
    elif 'time' not in data[0] and 'end' not in data[0]:
        print(data[0])
        if data[0] in xlsx_column:
            print(data[0])
            for j in range(3):
                work_sheet.cell(sumup_row, int(xlsx_column[data[0]][j])).value = data[j+1]
                work_sheet.cell(sumup_row, int(xlsx_column[data[0]][j])).data_type = 'float'
work_book.save(folder + xlsx_file)
