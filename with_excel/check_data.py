# 毎日の測定データは正弦波かどうかを確認するプログラムである。
# フォルダfolderの中にあるcsvファイルをすべて処理する

# 動作:

import os
import openpyxl
import csv
import with_excel.create_graph as cg 


def check(folder):
    if not os.path.exists(folder + 'excel'):
        os.mkdir(folder + 'excel')

    if not os.path.exists(folder + 'result'):
        os.mkdir(folder + 'result')

    sumup_file = open(folder + 'result/sumup.csv', 'a')
    # すべてのcsvファイル名前をfile_listに入れる
    file_list = []

    for file in os.listdir(folder):
        if '.csv' in file and '-' in file:
            file_list.append(file)

    for csv_name in file_list:
        wb = openpyxl.Workbook()
        ws = wb.active
        csv_file = folder + csv_name
        xlsx_file = folder + 'excel/' + csv_name.split('.')[0] + '.xlsx'

        with open(csv_file, 'r') as f:
            for row in csv.reader(f):
                ws.append(row)

        data = []

        with open(csv_file) as f:
            for row in f:
                data.append(row)

        data_address = []
        start_point = 0


        sumup_file.write(csv_name.split('-n')[0] + ',,,\n')
        sumup_file.write('time,henkou(%),unryou(%),ch1,\n')

        for i in range(data.__len__()):
            if 'henko(LUX)' in data[i]:
                start_point = i + 2

            ch1_list = []
            if 'henkoudo' in data[i]:
                for row in range(start_point,i+1):
                    ch1_list.append(data[row].split(',')[3])
                min_ch1 = 99999
                if ch1_list == ['']:
                    min_ch1 = 0
                else:
                    for j in ch1_list:
                        if j != '':
                            if int(j) < min_ch1 :
                                min_ch1 = int(j)

                data_address.append([start_point, i])
                data_split = data[i + 1].split(',')
                sumup_file.write(data[start_point-3].split(',')[3].split('\n')[0] + ',' + str(data_split[5]) + ',' + str(
                    data_split[6].split('\n')[0] + ','+str(min_ch1)+',\n'))

        csv_name_split = csv_name.split('-')

        # 　sizen(LUX),henko(LUX)…… から max(LUX),min(LUX)…… までの数値を float　にする
        if csv_name_split[1] == '10':
            for k in data_address:
                for i in range(k[0], k[1] + 1):
                    for j in range(1, 3):
                        ws.cell(i, j).data_type = 'float'

        #　ここは、11月3日午後から、ch0とch1のデータを記録始めたため
        elif csv_name_split[1] == '11' and csv_name_split[2] == '01' or csv_name_split[2] == '02' or csv_name_split[2] == '03':
            for k in data_address:
                for i in range(k[0], k[1] + 1):
                    for j in range(1, 3):
                        ws.cell(i, j).data_type = 'float'
        # ここは、ch0とch1のデータを記録後の操作
        else:
            for k in data_address:
                for i in range(k[0], k[1] + 1):
                    for j in range(1, 8):
                        ws.cell(i, j).data_type = 'float'



        # max(LUX),min(LUX)下の一行をfloatにする
        for k in data_address:
            for i in range(1, 8):
                ws.cell(k[1] + 2, i).data_type = 'float'

        for i in range(36):
            ws.cell(i + 3, 8).value = i * 5

        wb.save(xlsx_file)

        graph = cg.create_graph(xlsx_file, 'Sheet')

        # エクセルファイル中グラフ作成
        for k in data_address:

            if k == data_address[0]:
                load_file = graph.load()
                print('load %s'%csv_name)

            graph.create_scatter(load_file, '角度', '照度', '偏光' + '(' + data[k[0] - 3].split(',')[3].split('\n')[0] + ')',
                                 'H' + str(k[0]),
                                 [[8, 3, 8, 39, 2, k[0], 2, k[1],'henko']])
            graph.create_scatter(load_file, '角度', '照度', 'ch1 and ch2', 'O' + str(k[0]),
                                 [[8, 3, 8, 39, 3, k[0], 3, k[1],'ch0'], [8, 3, 8, 39, 4, k[0], 4, k[1],'ch1']])
            graph.create_scatter(load_file, '角度', '照度', 'lux1 and lux2', 'H' + str(k[0] + 17),
                                 [[8, 3, 8, 39, 5, k[0], 5, k[1],'lux1'], [8, 3, 8, 39, 6, k[0], 6, k[1],'lux2']])

            if k == data_address[data_address.__len__() - 1]:
                graph.save(load_file)
                print('save')




    sumup_file.write('end')

    print('完了')

if __name__ == '__main__':
    # 自分のフォルダを選択してください
    folder = 'C://Users/zyxx/Desktop/test_csv/daily_data/'
    # folder = 'F://daily/'
    check(folder)
