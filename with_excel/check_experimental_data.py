# 実験空得られたデータのグラフを作成するプログラムである。
# フォルダ(folder)の中にあるcsvファイルをすべて処理する


# 追加すべきもの
#   最大値、最小値,偏光度の計算 　(完了)
#   データをまとめるところ
#   X21から、まとめるデータを入力すること　(完了)

# 動作:書いてない

import os
import openpyxl
import csv
import with_excel.create_graph as cg

experimental_date = '2023年/01月/30日'

data_title = {1: '振らない', 2: '振る', 3: '振る', 4: '振る', 5: '振らない', 6: '振る',7:'振る',8:'振る',9:'1',10:'1',11:'1',12:'1',13:'1',14:'1',15:'1',16:'1',17:'1',18:
              1,19:'1',20:'1',21:'1',22:'1',23:'1',24:'1',25:'1'}
sheet_title = {'empty': 0, 'RX_OX': 1, 'torumarin_0': 2, 'torumarin_1': 3, 'torumarin_3': 4,
               'kokuen': 5, 'komugiko': 6, 'sugikafun': 7, 'tomato': 8,'none':9}
sheet_name_Japanses = {'empty': '空き', 'RX_OX': 'RX OX', 'torumarin_0': 'トルマリン(0.8)', 'torumarin_1': 'トルマリン(1.8)', 'torumarin_3': 'トルマリン(3)',
               'kokuen': '黒鉛', 'komugiko': '小麦粉', 'sugikafun': '花粉', 'tomato': 'トマト','none':'none'}
fill_color = openpyxl.styles.PatternFill('solid', fgColor='FFFF00')
border_set = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thick', color=openpyxl.styles.colors.BLACK),
                                    right=openpyxl.styles.Side(style='thick', color=openpyxl.styles.colors.BLACK),
                                    top=openpyxl.styles.Side(style='thick', color=openpyxl.styles.colors.BLACK),
                                    bottom=openpyxl.styles.Side(style='thick', color=openpyxl.styles.colors.BLACK))
global sumup_start_point_x
global sumup_start_point_y

sumup_start_point_x = 20
sumup_start_point_y = 24
sumup_number = 1
sumup_title = {0: '', 1: 'max(Lux)', 2: 'min(Lux)', 3: 'henko(%)'}


def check(folder):
    sheet_no = 0
    if not os.path.exists(folder):
        os.mkdir(folder)
    if not os.path.exists(folder + 'excel'):
        os.mkdir(folder + 'excel')

    # すべてのcsvファイル名前をfile_listに入れる
    file_list = []

    for file in os.listdir(folder):
        if '.csv' in file and '-e' in file:
            file_list.append(file)

    xlsx_file = folder + 'excel/' + file_list[0].split('-e-')[0] + '-e.xlsx'

    wb = openpyxl.Workbook()

    for csv_name in file_list:
        sheet_name = csv_name.split('-e-')[1].split('.')[0]
        sheet_no = sheet_title[sheet_name]
        ws = wb.create_sheet(sheet_name, sheet_no)
        print(sheet_no)
        if csv_name == file_list[0]:
            ws_sumup = wb.create_sheet('sumup', 9)
        else:
            ws_sumup = wb['sumup']
        csv_file = folder + csv_name

        with open(csv_file, 'r') as f:
            for row in csv.reader(f):
                ws.append(row)

        data = []
        with open(csv_file) as f:
            for row in f:
                data.append(row)

        data_address = []
        check = 0
        start_point = 0

        for i in range(data.__len__()):
            if 'CH0' in data[i]:
                start_point = i + 2
                if check == 0:
                    check = 1
            if '180' in data[i].split(',')[0]:
                check = 0
                data_address.append([start_point, i + 1])

        for k in data_address:
            for i in range(k[0], k[1] + 1):
                for j in range(1, 7):
                    ws.cell(i, j).data_type = 'float'

        for i in range(4):
            ws.cell(sumup_start_point_x, sumup_start_point_y + i).value = sumup_title[i]

        for k in data_address:
            ws.cell(k[0] - 2, 11).value = 'max(Lux)'
            ws.cell(k[0] - 2, 12).value = 'min(Lux)'
            ws.cell(k[0] - 2, 13).value = 'henko(%)'
            ws.cell(k[0] - 1, 11).value = '=MAX(B%s:B%s)' % (str(k[0]), str(k[1]))
            ws.cell(k[0] - 1, 12).value = '=MIN(B%s:B%s)' % (str(k[0]), str(k[1]))
            ws.cell(k[0] - 1, 13).value = '=(K%s-L%s)/(K%s+L%s)*100' % (
                str(k[0] - 1), str(k[0] - 1), str(k[0] - 1), str(k[0] - 1))

            for i in range(2):
                for j in range(3):
                    ws.cell(k[0] + i - 2, 11 + j).fill = fill_color
                    ws.cell(k[0] + i - 2, 11 + j).border = border_set

        graph = cg.create_graph(xlsx_file, sheet_name)

        data_list = []

        for k in data_address:
            if k == data_address[0]:
                load_file = graph.load_already(workbook=wb, worksheet=ws)
                print('load')
            # 散布図を作る
            graph.create_scatter(load_file, '角度(°)', '照度(Lux)',
                                 '偏光' + '(' + data[k[0] - 3].split(',')[3].split('\n')[0] + ')',
                                 'H' + str(k[0]),
                                 [[1, k[0], 1, k[1], 2, k[0], 2, k[1], 'henko']])
            graph.create_scatter(load_file, '角度(°)', '照度(Lux)', 'ch0 and ch1', 'O' + str(k[0]),
                                 [[1, k[0], 1, k[1], 3, k[0], 3, k[1], 'ch0'],
                                  [1, k[0], 1, k[1], 4, k[0], 4, k[1], 'ch1']])
            graph.create_scatter(load_file, '角度(°)', '照度(Lux)', 'lux1 and lux2', 'H' + str(k[0] + 17),
                                 [[1, k[0], 1, k[1], 5, k[0], 5, k[1], 'lux1'],
                                  [1, k[0], 1, k[1], 6, k[0], 6, k[1], 'lux2']])

            if k == data_address[0]:
                number = 1
                for i in data_address:
                    # CH1をグラフする
                    # data_list.append([1, i[0], 1, i[1], 4, i[0], 4, i[1],date_title])

                    # henkoをグラフにする
                    data_list.append([1, i[0], 1, i[1], 2, i[0], 2, i[1], data_title[number]])
                    number += 1

                graph.create_scatter(load_file, '角度(°)', '照度(Lux)',
                                     '偏光', 'P20', data_list)
        # ws_sumup :　統計用シート　ws : 今のシート
        sumup_no = sheet_no + 1
        # A=1,B=2,C=3,D=4,E=5,F=6,G=7,H=8,I=9,J=10,K=11,L=12,M=13,N=14
        sumup_sheet_i = 0
        sumup_sheet_start_point_x = 1
        if sumup_no % 3 == 0:
            sumup_sheet_start_point_y = 11
        else:
            sumup_sheet_start_point_y = 1 + (sumup_no % 3 - 1) * 5
        while ws_sumup.cell(sumup_sheet_start_point_x, sumup_sheet_start_point_y).value is not None and \
                ws_sumup.cell(sumup_sheet_start_point_x+1, sumup_sheet_start_point_y).value is not None:
            sumup_sheet_start_point_x += 1
        if sumup_sheet_start_point_x != 1:
            sumup_sheet_start_point_x += 1
        ws_sumup.cell(sumup_sheet_start_point_x, sumup_sheet_start_point_y).value = sheet_name_Japanses[sheet_name]
        ws_sumup.cell(sumup_sheet_start_point_x, sumup_sheet_start_point_y + 1).value = 'max(Lux)'
        ws_sumup.cell(sumup_sheet_start_point_x, sumup_sheet_start_point_y + 2).value = 'min(Lux)'
        ws_sumup.cell(sumup_sheet_start_point_x, sumup_sheet_start_point_y + 3).value = '偏光度(%)'

        while (ws.cell(sumup_sheet_i * 64 + 1, 11)).value is not None:
            ws_sumup.cell(sumup_sheet_start_point_x + sumup_sheet_i + 1,
                          sumup_sheet_start_point_y).value = '第' + str(sumup_sheet_i + 1) + '回'
            ws_sumup.cell(sumup_sheet_start_point_x + sumup_sheet_i + 1,
                          sumup_sheet_start_point_y + 1).value = '=' + sheet_name + '!' + 'K' + str(
                sumup_sheet_i * 64 + 2)
            ws_sumup.cell(sumup_sheet_start_point_x + sumup_sheet_i + 1,
                          sumup_sheet_start_point_y + 2).value = '=' + sheet_name + '!' + 'L' + str(
                sumup_sheet_i * 64 + 2)
            ws_sumup.cell(sumup_sheet_start_point_x + sumup_sheet_i + 1,
                          sumup_sheet_start_point_y + 3).value = '=' + sheet_name + '!' + 'M' + str(
                sumup_sheet_i * 64 + 2)
            sumup_sheet_i += 1
        ws_sumup.cell(sumup_sheet_start_point_x + sumup_sheet_i + 1,
                      sumup_sheet_start_point_y ).value = '平均(振らない)'
        ws_sumup.cell(sumup_sheet_start_point_x + sumup_sheet_i + 2,
                      sumup_sheet_start_point_y ).value = '平均(振る)'
        ws_sumup.cell(sumup_sheet_start_point_x + sumup_sheet_i + 3,
                      sumup_sheet_start_point_y ).value = ''

    wb.save(xlsx_file)
    print('save')

    print(data)


if __name__ == '__main__':
    folder = 'C://Users/zyxx/Desktop/test_csv/experimental_data/'
    os.system('rclone sync gdrive_taka:偏光測定器_実験データ/%s %s' % (experimental_date, folder))
    check(folder)
