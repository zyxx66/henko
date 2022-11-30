# 古いデータを含めて処理するプログラムである
# 古いデータとは：  ch0,ch1,lux0,lux1,unryo を記載されていないcsvファイル (2022-11-02以前のデータ)
# -nが付いてるcsvファイル：雲量以外のデータ
# -nが付いてないcsvファイル：雲量データ
#

# 毎日の測定データは正弦波かどうかを確認するプログラムである。
# フォルダfolderの中にあるcsvファイルをすべて処理する

# 動作:

import os
import openpyxl
import csv
import with_excel.create_graph as cg
import sum_up
import rclone_method as rc


def check(folder):
    if not os.path.exists(folder + 'excel'):
        os.mkdir(folder + 'excel')

    if not os.path.exists(folder + 'result'):
        os.mkdir(folder + 'result')

    sumup_file = open(folder + 'result/sumup.csv', 'a')
    # すべてのcsvファイル名前をfile_listに入れる
    file_list = []
    unryo_file_list = []
    experimental_file_list = []

    for file in os.listdir(folder):
        # 偏光測定器データ(雲量以外)
        if '.csv' in file and '-n' in file:
            file_list.append(file)
        # 雲量データ
        if '.csv' in file and '-n' not in file:
            unryo_file_list.append(file)
        # 実験データ
        if '.csv' in file and '-e' in file:
            experimental_file_list.append(file)
    # 偏光測定器データが存在すれば
    if file_list.__len__() != 0:
        for csv_name in file_list:
            wb = openpyxl.Workbook()
            ws = wb.active
            csv_file = folder + csv_name
            xlsx_file = folder + 'excel/' + csv_name.split('.')[0] + '.xlsx'

            with open(csv_file, 'r') as f:
                for row in csv.reader(f):
                    ws.append(row)

            data = []
            with open(csv_file) as f:
                for row in f:
                    data.append(row)

            data_address = []
            check = 0
            start_point = 0


            sumup_file.write(csv_name.split('-n')[0] + ',,,\n')
            sumup_file.write('time,henkou(%),unryou(%),,\n')
            for i in range(data.__len__()):
                if 'ch0' in data[i]:
                    start_point = i + 2
                    if check == 0:
                        check = 1
                if 'henkoudo' in data[i]:
                    check = 0
                    data_address.append([start_point, i])
                if 'henkoudo' in data[i]:
                    data_split = data[i + 1].split(',')
                    sumup_file.write(data[i - 39].split(',')[3].split('\n')[0] + ',' + str(data_split[5]) + ',' + str(
                        data_split[6].split('\n')[0] + ',,\n'))

            for k in data_address:
                for i in range(k[0], k[1] + 1):
                    for j in range(1, 8):
                        ws.cell(i, j).data_type = 'float'

            for k in data_address:
                for i in range(1, 8):
                    ws.cell(k[1] + 2, i).data_type = 'float'

            for i in range(36):
                ws.cell(i + 3, 8).value = i * 5

            wb.save(xlsx_file)

            graph = cg.create_graph(xlsx_file, 'Sheet')

            for k in data_address:

                if k == data_address[0]:
                    load_file = graph.load()
                    print('load')

                graph.create_scatter(load_file, '角度', '照度', '偏光' + '(' + data[k[0] - 3].split(',')[3].split('\n')[0] + ')',
                                     'H' + str(k[0]),
                                     [8, 3, 8, 39, 2, k[0], 2, k[1]])
                graph.create_scatter(load_file, '角度', '照度', 'ch1 and ch2', 'O' + str(k[0]),
                                     [8, 3, 8, 39, 3, k[0], 3, k[1]], [8, 3, 8, 39, 4, k[0], 4, k[1]])
                graph.create_scatter(load_file, '角度', '照度', 'lux1 and lux2', 'H' + str(k[0] + 17),
                                     [8, 3, 8, 39, 5, k[0], 5, k[1]], [8, 3, 8, 39, 6, k[0], 6, k[1]])

                if k == data_address[data_address.__len__() - 1]:
                    graph.save(load_file)
                    print('save')




    sumup_file.write('end')

    print(data)


if __name__ == '__main__':
    folder = 'C://Users/zyxx/Desktop/test_csv/'
    check(folder)
    sum_up.sumup(folder)
